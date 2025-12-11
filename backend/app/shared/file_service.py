"""Generic FileService for handling file uploads and storage by ID."""
import uuid
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from typing import BinaryIO

import pandas as pd
from fastapi import UploadFile, HTTPException

from app.core.config import settings


class FileService:
    """Service for managing temporary file storage and retrieval."""
    
    def __init__(self):
        self.temp_dir = settings.temp_files_dir
        self.temp_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_file_id(self) -> str:
        """Generate a unique file ID."""
        return str(uuid.uuid4())
    
    async def save_upload(self, upload_file: UploadFile) -> tuple[str, Path]:
        """
        Save an uploaded file and return its file_id and path.
        
        Args:
            upload_file: The uploaded file from FastAPI
            
        Returns:
            Tuple of (file_id, file_path)
            
        Raises:
            HTTPException: If file extension is not allowed or file is too large
        """
        # Validate file extension
        file_ext = Path(upload_file.filename).suffix.lower()
        if file_ext not in settings.allowed_extensions:
            raise HTTPException(
                status_code=400,
                detail=f"File extension {file_ext} not allowed. Allowed: {settings.allowed_extensions}"
            )
        
        # Generate unique file ID
        file_id = self.generate_file_id()
        file_path = self.temp_dir / f"{file_id}{file_ext}"
        
        # Save file
        try:
            with file_path.open("wb") as buffer:
                shutil.copyfileobj(upload_file.file, buffer)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to save file: {str(e)}")
        
        return file_id, file_path
    
    def get_file_path(self, file_id: str) -> Path:
        """
        Get the file path for a given file_id.
        
        Args:
            file_id: The unique file identifier
            
        Returns:
            Path to the file
            
        Raises:
            HTTPException: If file not found
        """
        # Search for file with any allowed extension
        for ext in settings.allowed_extensions:
            file_path = self.temp_dir / f"{file_id}{ext}"
            if file_path.exists():
                return file_path
        
        raise HTTPException(status_code=404, detail=f"File with ID {file_id} not found")
    
    def load_excel(self, file_id: str) -> pd.DataFrame:
        """
        Load an Excel file into a pandas DataFrame.
        
        Args:
            file_id: The unique file identifier
            
        Returns:
            pandas DataFrame containing the Excel data
            
        Raises:
            HTTPException: If file not found or cannot be loaded
        """
        file_path = self.get_file_path(file_id)
        
        try:
            # Read Excel file
            df = pd.read_excel(file_path, engine='openpyxl')
            return df
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to load Excel file: {str(e)}"
            )
    
    
    def get_excel_preview(self, file_id: str, max_rows: int = 50) -> tuple[pd.DataFrame, int]:
        """
        Get preview DataFrame and total row count efficiently.
        
        Args:
            file_id: The unique file identifier
            max_rows: Maximum rows to read
            
        Returns:
            Tuple of (preview_df, total_rows)
        """
        file_path = self.get_file_path(file_id)
        
        try:
            # 1. Get total rows efficiently using openpyxl read-only mode
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            # max_row in read_only might be None if not valid metadata, but usually fine for saved files
            total_rows = ws.max_row if ws.max_row is not None else 0
            # Adjust for header row (assuming 1 header row)
            total_rows = max(0, total_rows - 1)
            wb.close()
            
            # 2. Read only preview rows using pandas
            df = pd.read_excel(file_path, engine='openpyxl', nrows=max_rows)
            
            return df, total_rows
            
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to load Excel preview: {str(e)}"
            )

    def save_dataframe(self, df: pd.DataFrame, original_file_id: str | None = None) -> str:

        """
        Save a pandas DataFrame to a new Excel file and return new file_id.
        
        Args:
            df: The DataFrame to save
            original_file_id: Optional original file ID to determine file extension
            
        Returns:
            new file_id for the saved file
            
        Raises:
            HTTPException: If save operation fails
        """
        # Generate new file ID
        new_file_id = self.generate_file_id()
        
        # Use .xlsx as default extension
        file_ext = ".xlsx"
        if original_file_id:
            try:
                original_path = self.get_file_path(original_file_id)
                file_ext = original_path.suffix
            except:
                pass  # Use default .xlsx
        
        file_path = self.temp_dir / f"{new_file_id}{file_ext}"
        
        # Save DataFrame to Excel
        try:
            df.to_excel(file_path, index=False, engine='openpyxl')
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to save Excel file: {str(e)}"
            )
        
        return new_file_id
    
    def cleanup_old_files(self, hours: int | None = None):
        """
        Remove files older than specified hours.
        
        Args:
            hours: Number of hours to retain files (default from settings)
        """
        if hours is None:
            hours = settings.file_retention_hours
        
        cutoff_time = datetime.now() - timedelta(hours=hours)
        
        for file_path in self.temp_dir.iterdir():
            if file_path.is_file():
                file_modified = datetime.fromtimestamp(file_path.stat().st_mtime)
                if file_modified < cutoff_time:
                    try:
                        file_path.unlink()
                    except Exception:
                        pass  # Ignore cleanup errors
    
    def delete_file(self, file_id: str) -> bool:
        """
        Delete a specific file by file_id.
        
        Args:
            file_id: The file identifier to delete
            
        Returns:
            True if deleted, False if not found
        """
        try:
            file_path = self.get_file_path(file_id)
            file_path.unlink()
            return True
        except HTTPException:
            return False
